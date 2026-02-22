# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  CONVERSOR EXCEL → JSON - CATÁLOGO DE PRODUTOS SISCOMEX (CATP API)        ║
║  Portal Único Siscomex - Portal Único de Comércio Exterior                 ║
║  Versão: 2.0                                                               ║
╚══════════════════════════════════════════════════════════════════════════════╝

Converte planilha Excel (.xlsx) para JSON compatível com a API do Catálogo de
Produtos do Portal Único Siscomex (CATP).

Suporta dois schemas da API:
- ProdutoIntegracaoRequestDTO (novos endpoints): POST /ext/produto/{cpfCnpjRaiz}
- ProdutoIntegracaoDTO (endpoint depreciado): POST /ext/produto (lote com seq)

Formatos de saída:
- api_post:  Inclusão via API (ProdutoIntegracaoRequestDTO, sem seq/cpfCnpjRaiz/situacao)
- api_put:   Atualização via API (ProdutoIntegracaoRequestDTO, sem seq/cpfCnpjRaiz/situacao)
- post:      Upload em lote no portal (ProdutoIntegracaoDTO, com seq)
- put:       Atualização em lote no portal (ProdutoIntegracaoDTO, com seq e codigo)
- completo:  Exportação completa (formato idêntico ao exportado pelo portal)

Genérico: funciona para qualquer empresa (CNPJ configurável).
"""

import json
import os
import sys
import re
import zipfile
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("=" * 70)
    print("ERRO: Biblioteca 'openpyxl' não encontrada.")
    print("Instale com: pip install openpyxl")
    print("=" * 70)
    sys.exit(1)


# ============================================================================
# CONSTANTES E VALIDAÇÕES DA API CATP
# ============================================================================

# Modalidades válidas
MODALIDADES_VALIDAS = ["IMPORTACAO", "EXPORTACAO"]

# Situações válidas (API aceita maiúsculas)
SITUACOES_VALIDAS = ["ATIVADO", "DESATIVADO", "RASCUNHO"]
# Mapeamento para normalizar valores de situação
SITUACAO_NORMALIZAR = {
    "ativado": "ATIVADO",
    "desativado": "DESATIVADO",
    "rascunho": "RASCUNHO",
}

# Tamanhos máximos de campos (conforme API)
MAX_DENOMINACAO = 120  # Swagger spec: Tamanho máximo: 120
MAX_DESCRICAO = 2000
MAX_NCM = 8
MAX_CODIGO_INTERNO = 60
MAX_VALOR_ATRIBUTO = 3000
MAX_CPF_CNPJ_RAIZ = 14  # CNPJ raiz = 8, CPF = 11, mas campo aceita até 14

# Campos read-only (gerados pelo servidor, NÃO enviar no POST)
CAMPOS_READ_ONLY = ["seq", "versao"]

# Campos obrigatórios para criação via API nova (cpfCnpjRaiz vai na URL)
CAMPOS_OBRIGATORIOS_API = ["descricao", "denominacao", "ncm", "modalidade"]

# Campos obrigatórios para criação via upload em lote (portal, endpoint depreciado)
CAMPOS_OBRIGATORIOS_POST = ["descricao", "denominacao", "ncm", "cpfCnpjRaiz", "modalidade"]

# Campos obrigatórios para atualização (PUT)
CAMPOS_OBRIGATORIOS_PUT = ["codigo", "descricao", "denominacao", "ncm", "cpfCnpjRaiz", "modalidade"]

# Colunas principais da planilha (ordem fixa)
COLUNAS_PRINCIPAIS = [
    "codigo",           # Código do produto (int, gerado pelo servidor no POST, obrigatório no PUT)
    "denominacao",      # Nome do produto (obrigatório, max 120)
    "descricao",        # Descrição detalhada (obrigatório, max 2000)
    "cpfCnpjRaiz",      # CNPJ raiz 8 dígitos (obrigatório)
    "situacao",          # ATIVADO/DESATIVADO (opcional, padrão ATIVADO)
    "modalidade",        # IMPORTACAO/EXPORTACAO (obrigatório)
    "ncm",               # Código NCM 8 dígitos (obrigatório)
    "codigosInterno",    # Códigos internos separados por ; (opcional)
]

# Mapeamento de atributos conhecidos para NCMs comuns (para labels amigáveis)
ATRIBUTOS_LABELS = {
    "ATT_14540": "Condição do Produto",
    "ATT_14545": "País de Origem (Código)",
    "ATT_14546": "Validade do Produto",
    "ATT_14547": "Produto Controlado",
    "ATT_14551": "Registro ANVISA",
    "ATT_14554": "Produto Perigoso",
    "ATT_14555": "Fabricante/Exportador",
    "ATT_14556": "Tipo de Embalagem",
    "ATT_14860": "Nome Comercial",
    "ATT_15120": "Composição/Material",
    "ATT_15121": "Modelo/Referência",
}


# ============================================================================
# CLASSE PRINCIPAL DE CONVERSÃO
# ============================================================================

class ConversorCatalogoSiscomex:
    """Converte planilha Excel para JSON no padrão CATP API Siscomex."""

    def __init__(self, auto_truncar=False):
        self.erros = []
        self.avisos = []
        self.produtos = []
        self.auto_truncar = auto_truncar  # Truncar campos que excedem o limite

    # ========================================================================
    # VALIDAÇÃO DE CAMPOS
    # ========================================================================

    def validar_ncm(self, ncm: str, linha: int) -> bool:
        """Valida formato do NCM (8 dígitos numéricos)."""
        ncm_limpo = str(ncm).strip().replace(".", "").replace("-", "").replace(" ", "")
        if not ncm_limpo.isdigit():
            self.erros.append(f"Linha {linha}: NCM '{ncm}' contém caracteres não numéricos.")
            return False
        if len(ncm_limpo) != 8:
            self.erros.append(f"Linha {linha}: NCM '{ncm}' deve ter exatamente 8 dígitos (tem {len(ncm_limpo)}).")
            return False
        return True

    def validar_modalidade(self, modalidade: str, linha: int) -> bool:
        """Valida modalidade (IMPORTACAO ou EXPORTACAO)."""
        if modalidade.upper() not in MODALIDADES_VALIDAS:
            self.erros.append(
                f"Linha {linha}: Modalidade '{modalidade}' inválida. "
                f"Valores aceitos: {', '.join(MODALIDADES_VALIDAS)}"
            )
            return False
        return True

    def normalizar_situacao(self, situacao: str) -> str:
        """Normaliza o valor de situação para maiúsculas conforme API."""
        if not situacao:
            return "ATIVADO"
        return SITUACAO_NORMALIZAR.get(situacao.strip().lower(), situacao.upper())

    def validar_situacao(self, situacao: str, linha: int) -> bool:
        """Valida situação do produto."""
        situacao_norm = self.normalizar_situacao(situacao) if situacao else "ATIVADO"
        if situacao_norm not in SITUACOES_VALIDAS:
            self.erros.append(
                f"Linha {linha}: Situação '{situacao}' inválida. "
                f"Valores aceitos: {', '.join(SITUACOES_VALIDAS)}"
            )
            return False
        return True

    def validar_cpf_cnpj_raiz(self, valor: str, linha: int) -> bool:
        """Valida CPF/CNPJ raiz (somente dígitos)."""
        valor_limpo = str(valor).strip().replace(".", "").replace("-", "").replace("/", "")
        if not valor_limpo.isdigit():
            self.erros.append(f"Linha {linha}: cpfCnpjRaiz '{valor}' deve conter apenas dígitos.")
            return False
        if len(valor_limpo) > MAX_CPF_CNPJ_RAIZ:
            self.erros.append(f"Linha {linha}: cpfCnpjRaiz '{valor}' excede {MAX_CPF_CNPJ_RAIZ} caracteres.")
            return False
        return True

    def validar_campo_obrigatorio(self, valor, campo: str, linha: int) -> bool:
        """Valida se campo obrigatório está preenchido."""
        if valor is None or str(valor).strip() == "":
            self.erros.append(f"Linha {linha}: Campo obrigatório '{campo}' está vazio.")
            return False
        return True

    def validar_tamanho(self, valor: str, campo: str, maximo: int, linha: int) -> bool:
        """Valida tamanho máximo de campo."""
        if valor and len(str(valor)) > maximo:
            self.erros.append(
                f"Linha {linha}: Campo '{campo}' excede {maximo} caracteres "
                f"(tem {len(str(valor))})."
            )
            return False
        return True

    # ========================================================================
    # LEITURA E PROCESSAMENTO DA PLANILHA
    # ========================================================================

    def ler_planilha(self, caminho_excel: str, defaults: dict = None) -> list:
        """Lê a planilha Excel e retorna lista de produtos.
        
        Args:
            caminho_excel: Caminho do arquivo .xlsx
            defaults: Dict com valores padrão para campos ausentes na planilha
                      Ex: {'cpfCnpjRaiz': '12345678', 'modalidade': 'IMPORTACAO'}
        """
        defaults = defaults or {}
        if not os.path.exists(caminho_excel):
            self.erros.append(f"Arquivo não encontrado: {caminho_excel}")
            return []

        print(f"\n📂 Lendo planilha: {caminho_excel}")

        try:
            wb = openpyxl.load_workbook(caminho_excel, data_only=True)
        except zipfile.BadZipFile:
            self.erros.append(
                "O arquivo não é um .xlsx válido. Provavelmente está no formato "
                "antigo .xls renomeado para .xlsx. Abra o arquivo no Excel e "
                "salve como 'Pasta de Trabalho do Excel (.xlsx)' usando Salvar Como."
            )
            return []
        except Exception as e:
            self.erros.append(f"Erro ao abrir planilha: {str(e)}")
            return []

        ws = wb.active

        # Ler cabeçalhos da primeira linha
        cabecalhos = []
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=1, column=col).value
            if valor is not None:
                cabecalhos.append(str(valor).strip())
            else:
                cabecalhos.append(f"COL_{col}")

        print(f"📋 Colunas encontradas: {len(cabecalhos)}")
        print(f"   {', '.join(cabecalhos[:10])}{'...' if len(cabecalhos) > 10 else ''}")

        # Identificar colunas de atributos (começa com ATT_)
        colunas_atributos_simples = {}       # {indice: codigo_atributo}
        colunas_atributos_multi = {}          # {indice: codigo_atributo}
        colunas_principais = {}               # {nome_campo: indice}

        for idx, cab in enumerate(cabecalhos):
            cab_upper = cab.upper().strip()

            # Verifica se é coluna de atributo multivalorado (sufixo _MULTI ou [MULTI])
            if cab_upper.startswith("ATT_") and ("_MULTI" in cab_upper or "[MULTI]" in cab_upper):
                codigo_att = re.match(r"(ATT_\d+)", cab_upper).group(1)
                colunas_atributos_multi[idx] = codigo_att
            elif cab_upper.startswith("ATT_"):
                # Atributo simples
                codigo_att = re.match(r"(ATT_\d+)", cab_upper).group(1)
                colunas_atributos_simples[idx] = codigo_att
            else:
                # Campo principal - mapear por nome
                for campo in COLUNAS_PRINCIPAIS:
                    if cab_upper == campo.upper() or cab_upper.replace(" ", "").replace("_", "") == campo.upper().replace("_", ""):
                        colunas_principais[campo] = idx
                        break
                else:
                    # Tentar mapeamentos alternativos comuns
                    mapa_alternativos = {
                        "CODIGO": "codigo",
                        "COD": "codigo",
                        "CÓDIGO": "codigo",
                        "DENOMINACAO": "denominacao",
                        "DENOMINAÇÃO": "denominacao",
                        "NOME": "denominacao",
                        "NOME_PRODUTO": "denominacao",
                        "NOME DO PRODUTO": "denominacao",
                        "TITULO": "denominacao",
                        "TÍTULO": "denominacao",
                        "PRODUTO": "denominacao",
                        "NOME PRODUTO": "denominacao",
                        "NOME COMERCIAL": "denominacao",
                        "DESCRICAO": "descricao",
                        "DESCRIÇÃO": "descricao",
                        "DESCRICAO_PRODUTO": "descricao",
                        "DESCRIÇÃO DO PRODUTO": "descricao",
                        "DESCRICAO DETALHADA": "descricao",
                        "CNPJ": "cpfCnpjRaiz",
                        "CNPJ_RAIZ": "cpfCnpjRaiz",
                        "CPF_CNPJ": "cpfCnpjRaiz",
                        "CPFCNPJRAIZ": "cpfCnpjRaiz",
                        "CPF/CNPJ RAIZ": "cpfCnpjRaiz",
                        "CNPJ RAIZ": "cpfCnpjRaiz",
                        "SITUACAO": "situacao",
                        "SITUAÇÃO": "situacao",
                        "STATUS": "situacao",
                        "ATIVO": "situacao",
                        "MODALIDADE": "modalidade",
                        "TIPO": "modalidade",
                        "TIPO OPERACAO": "modalidade",
                        "TIPO OPERAÇÃO": "modalidade",
                        "NCM": "ncm",
                        "CODIGO_NCM": "ncm",
                        "COD_NCM": "ncm",
                        "NCM/SH": "ncm",
                        "CLASSIFICACAO FISCAL": "ncm",
                        "CLASSIFICAÇÃO FISCAL": "ncm",
                        "CÓDIGOS INTERNOS": "codigosInterno",
                        "CODIGOS_INTERNO": "codigosInterno",
                        "CODIGOSINTERNO": "codigosInterno",
                        "CÓDIGOS INTERNO": "codigosInterno",
                        "CODIGO_INTERNO": "codigosInterno",
                        "COD_INTERNO": "codigosInterno",
                        "CODIGOS INTERNOS": "codigosInterno",
                        "CODIGO DE BARRAS": "codigosInterno",
                        "CÓDIGO DE BARRAS": "codigosInterno",
                        "COD BARRAS": "codigosInterno",
                        "EAN": "codigosInterno",
                        "GTIN": "codigosInterno",
                        "COD DE FABRICA": "codigosInterno",
                        "CÓD DE FÁBRICA": "codigosInterno",
                        "CODIGO DE FABRICA": "codigosInterno",
                        "REFERÊNCIA DO FORNECEDOR": "codigosInterno",
                        "REFERENCIA DO FORNECEDOR": "codigosInterno",
                        "REF FORNECEDOR": "codigosInterno",
                    }
                    cab_normalizado = cab_upper.replace(" ", "").replace("_", "").replace("-", "")
                    for chave, campo in mapa_alternativos.items():
                        chave_norm = chave.replace(" ", "").replace("_", "").replace("-", "")
                        if cab_normalizado == chave_norm:
                            colunas_principais[campo] = idx
                            break

        print(f"\n🔍 Mapeamento de colunas:")
        print(f"   Campos principais: {len(colunas_principais)}")
        for campo, idx in sorted(colunas_principais.items(), key=lambda x: x[1]):
            print(f"     Col {idx+1} ({cabecalhos[idx]}) → {campo}")
        print(f"   Atributos simples: {len(colunas_atributos_simples)}")
        for idx, att in sorted(colunas_atributos_simples.items()):
            label = ATRIBUTOS_LABELS.get(att, att)
            print(f"     Col {idx+1} ({cabecalhos[idx]}) → {att} ({label})")
        print(f"   Atributos multivalorados: {len(colunas_atributos_multi)}")
        for idx, att in sorted(colunas_atributos_multi.items()):
            label = ATRIBUTOS_LABELS.get(att, att)
            print(f"     Col {idx+1} ({cabecalhos[idx]}) → {att} ({label})")

        # Verificar campos obrigatórios (aceitar defaults para os que faltam)
        campos_faltando = []
        campos_usando_default = []
        for campo in CAMPOS_OBRIGATORIOS_POST:
            if campo not in colunas_principais:
                if campo in defaults and defaults[campo]:
                    campos_usando_default.append(f"{campo}='{defaults[campo]}'")
                else:
                    campos_faltando.append(campo)

        if campos_usando_default:
            self.avisos.append(
                f"Campos preenchidos com valor padrão: {', '.join(campos_usando_default)}"
            )

        if campos_faltando:
            dicas = []
            for c in campos_faltando:
                if c == 'denominacao':
                    dicas.append("'denominacao' (ou Titulo, Nome do Produto)")
                elif c == 'cpfCnpjRaiz':
                    dicas.append("'cpfCnpjRaiz' (CNPJ raiz 8 dígitos) — preencha o campo CNPJ Raiz no site")
                elif c == 'modalidade':
                    dicas.append("'modalidade' (IMPORTACAO/EXPORTACAO) — selecione a Modalidade no site")
                else:
                    dicas.append(f"'{c}'")
            self.erros.append(
                f"Colunas obrigatórias não encontradas: {', '.join(dicas)}. "
                f"Verifique os cabeçalhos ou preencha os valores padrão no site."
            )
            return []

        # Processar cada linha de dados (a partir da linha 2)
        produtos = []
        for row in range(2, ws.max_row + 1):
            # Verificar se a linha está vazia (checar pelo menos algum campo preenchido)
            linha_vazia = True
            for col_idx in range(len(cabecalhos)):
                val = ws.cell(row=row, column=col_idx + 1).value
                if val is not None and str(val).strip() != "":
                    linha_vazia = False
                    break
            if linha_vazia:
                continue

            produto = self._processar_linha(
                ws, row, cabecalhos,
                colunas_principais,
                colunas_atributos_simples,
                colunas_atributos_multi,
                defaults
            )
            if produto:
                produtos.append(produto)

        print(f"\n✅ {len(produtos)} produtos lidos com sucesso.")
        wb.close()
        return produtos

    def _processar_linha(self, ws, row, cabecalhos, cols_principais, cols_att_simples, cols_att_multi, defaults=None) -> dict:
        """Processa uma linha da planilha e retorna um dicionário de produto."""
        defaults = defaults or {}
        produto = {}
        linha_valida = True

        # 1. Campos principais
        for campo, idx in cols_principais.items():
            valor_celula = ws.cell(row=row, column=idx + 1).value

            if valor_celula is None:
                valor = ""
            else:
                valor = str(valor_celula).strip()

            # Limpeza e normalização
            if campo == "ncm":
                valor = valor.replace(".", "").replace("-", "").replace(" ", "")
                # Se veio como número float (ex: 90211010.0), remover .0
                if valor.endswith(".0"):
                    valor = valor[:-2]
                # Preencher zeros à esquerda se necessário
                valor = valor.zfill(8)

            elif campo == "cpfCnpjRaiz":
                valor = valor.replace(".", "").replace("-", "").replace("/", "").replace(" ", "")
                if valor.endswith(".0"):
                    valor = valor[:-2]

            elif campo == "modalidade":
                valor = valor.upper().strip()
                # Normalizar variações
                if valor in ["IMP", "IMPORT", "IMPORTAÇÃO", "IMPORTAÇAO"]:
                    valor = "IMPORTACAO"
                elif valor in ["EXP", "EXPORT", "EXPORTAÇÃO", "EXPORTAÇAO"]:
                    valor = "EXPORTACAO"

            elif campo == "situacao":
                if valor == "":
                    valor = "ATIVADO"  # Padrão (maiúscula conforme API)
                # Normalizar para MAIÚSCULAS conforme Swagger
                valor_lower = valor.lower()
                if valor_lower in ["ativo", "ativado", "sim", "s", "1", "true", "yes"]:
                    valor = "ATIVADO"
                elif valor_lower in ["inativo", "desativado", "não", "nao", "n", "0", "false", "no"]:
                    valor = "DESATIVADO"
                elif valor_lower == "rascunho":
                    valor = "RASCUNHO"
                else:
                    valor = valor.upper()  # Qualquer outro valor, forçar uppercase

            elif campo == "codigo":
                if valor and valor != "":
                    try:
                        valor = int(float(valor))
                    except (ValueError, TypeError):
                        pass

            elif campo == "codigosInterno":
                # Não processar aqui, será tratado separadamente
                pass

            produto[campo] = valor

        # 1b. Aplicar defaults para campos que não estão na planilha
        for campo_default, valor_default in defaults.items():
            if campo_default not in cols_principais:
                produto[campo_default] = str(valor_default).strip()

        # Se não tem denominacao mas tem descricao, usar descricao como denominacao
        if not produto.get('denominacao') and produto.get('descricao'):
            produto['denominacao'] = produto['descricao'][:MAX_DENOMINACAO]
            self.avisos.append(
                f"Linha {row}: 'denominacao' ausente, usando os primeiros "
                f"{MAX_DENOMINACAO} caracteres da 'descricao'."
            )

        # 2. Validações
        for campo in CAMPOS_OBRIGATORIOS_POST:
            if not self.validar_campo_obrigatorio(produto.get(campo), campo, row):
                linha_valida = False

        if produto.get("ncm") and not self.validar_ncm(produto["ncm"], row):
            linha_valida = False

        if produto.get("modalidade") and not self.validar_modalidade(produto["modalidade"], row):
            linha_valida = False

        if produto.get("situacao") and not self.validar_situacao(produto["situacao"], row):
            linha_valida = False

        if produto.get("cpfCnpjRaiz") and not self.validar_cpf_cnpj_raiz(produto["cpfCnpjRaiz"], row):
            linha_valida = False

        # Truncamento automático ou validação de tamanho
        if produto.get("denominacao") and len(str(produto["denominacao"])) > MAX_DENOMINACAO:
            if self.auto_truncar:
                original_len = len(produto["denominacao"])
                produto["denominacao"] = produto["denominacao"][:MAX_DENOMINACAO]
                self.avisos.append(
                    f"Linha {row}: 'denominacao' truncada de {original_len} para {MAX_DENOMINACAO} caracteres."
                )
            else:
                self.erros.append(
                    f"Linha {row}: Campo 'denominacao' excede {MAX_DENOMINACAO} caracteres "
                    f"(tem {len(produto['denominacao'])}). Ative 'Truncar automaticamente' para cortar."
                )
                linha_valida = False

        if produto.get("descricao") and len(str(produto["descricao"])) > MAX_DESCRICAO:
            if self.auto_truncar:
                original_len = len(produto["descricao"])
                produto["descricao"] = produto["descricao"][:MAX_DESCRICAO]
                self.avisos.append(
                    f"Linha {row}: 'descricao' truncada de {original_len} para {MAX_DESCRICAO} caracteres."
                )
            else:
                self.erros.append(
                    f"Linha {row}: Campo 'descricao' excede {MAX_DESCRICAO} caracteres "
                    f"(tem {len(produto['descricao'])}). Ative 'Truncar automaticamente' para cortar."
                )
                linha_valida = False

        if not linha_valida:
            return None

        # 3. Processar códigos internos (separados por ; ou ,)
        cod_internos_raw = produto.get("codigosInterno", "")
        if cod_internos_raw and str(cod_internos_raw).strip():
            # Suporta separadores: ; , | ou nova linha
            separador = re.compile(r'[;,|\n]+')
            codigos = [c.strip() for c in separador.split(str(cod_internos_raw)) if c.strip()]
            # Validar tamanho individual
            for cod in codigos:
                if len(cod) > MAX_CODIGO_INTERNO:
                    self.avisos.append(
                        f"Linha {row}: Código interno '{cod[:30]}...' excede {MAX_CODIGO_INTERNO} caracteres."
                    )
            produto["codigosInterno"] = codigos
        else:
            produto["codigosInterno"] = []

        # 4. Processar atributos simples
        atributos = []
        for idx, codigo_att in cols_att_simples.items():
            valor_celula = ws.cell(row=row, column=idx + 1).value
            if valor_celula is not None and str(valor_celula).strip() != "":
                valor_str = str(valor_celula).strip()
                # Tratar booleanos
                if isinstance(valor_celula, bool):
                    valor_str = "true" if valor_celula else "false"
                elif valor_str.upper() in ["TRUE", "VERDADEIRO", "SIM"]:
                    valor_str = "true"
                elif valor_str.upper() in ["FALSE", "FALSO", "NÃO", "NAO"]:
                    valor_str = "false"
                # Remover .0 de números inteiros
                if valor_str.endswith(".0"):
                    try:
                        float(valor_str)
                        valor_str = valor_str[:-2]
                    except ValueError:
                        pass

                # Zero-padding para atributos com código de domínio de 2 dígitos
                # ATT_14540 (Estágio de Fabricação): "1" → "01"
                if codigo_att == 'ATT_14540' and valor_str.isdigit() and len(valor_str) == 1:
                    valor_str = valor_str.zfill(2)

                atributos.append({
                    "atributo": codigo_att,
                    "valor": valor_str
                })

        produto["atributos"] = atributos

        # 5. Processar atributos multivalorados
        atributos_multi = []
        for idx, codigo_att in cols_att_multi.items():
            valor_celula = ws.cell(row=row, column=idx + 1).value
            if valor_celula is not None and str(valor_celula).strip() != "":
                valor_str = str(valor_celula).strip()
                # Valores separados por ; ou ,
                separador = re.compile(r'[;,|\n]+')
                valores = [v.strip() for v in separador.split(valor_str) if v.strip()]
                # Remover .0 de cada valor
                valores_limpos = []
                for v in valores:
                    if v.endswith(".0"):
                        try:
                            float(v)
                            v = v[:-2]
                        except ValueError:
                            pass
                    valores_limpos.append(v)

                if valores_limpos:
                    atributos_multi.append({
                        "atributo": codigo_att,
                        "valores": valores_limpos
                    })

        produto["atributosMultivalorados"] = atributos_multi

        # 6. Arrays vazios para compostos (preenchidos manualmente se necessário)
        produto["atributosCompostos"] = []
        produto["atributosCompostosMultivalorados"] = []

        return produto

    # ========================================================================
    # GERAÇÃO DE JSON
    # ========================================================================

    def gerar_json_post(self, produtos: list) -> list:
        """
        Gera JSON para POST (inclusão de novos produtos via upload no portal).
        Usa o schema ProdutoIntegracaoDTO que requer 'seq'.
        Remove campos read-only: versao, codigo.
        NÃO inclui versao nem codigo para que o portal crie novos produtos.
        Ordem dos campos segue o padrão do portal.
        """
        resultado = []
        for seq, produto in enumerate(produtos, 1):
            item = {}
            # Ordem: seq, descricao, denominacao, cpfCnpjRaiz, situacao,
            #         modalidade, ncm, atributos..., codigosInterno
            item["seq"] = seq
            item["descricao"] = produto.get("descricao", "")
            item["denominacao"] = produto.get("denominacao", "")
            item["cpfCnpjRaiz"] = produto.get("cpfCnpjRaiz", "")
            situacao = produto.get("situacao", "ATIVADO")
            item["situacao"] = self.normalizar_situacao(situacao)
            item["modalidade"] = produto.get("modalidade", "")
            item["ncm"] = produto.get("ncm", "")

            # Atributos
            item["atributos"] = produto.get("atributos", [])
            item["atributosMultivalorados"] = produto.get("atributosMultivalorados", [])
            item["atributosCompostos"] = produto.get("atributosCompostos", [])
            item["atributosCompostosMultivalorados"] = produto.get("atributosCompostosMultivalorados", [])

            # Códigos internos
            item["codigosInterno"] = produto.get("codigosInterno", [])

            resultado.append(item)

        return resultado

    def gerar_json_put(self, produtos: list) -> list:
        """
        Gera JSON para PUT (atualização/nova versão de produtos existentes).
        Inclui 'seq' e 'codigo' no body. Remove versao (nova versão é criada pelo servidor).
        Usa ProdutoIntegracaoDTO para upload em lote pelo portal.
        """
        resultado = []
        for seq, produto in enumerate(produtos, 1):
            codigo = produto.get("codigo")
            if not codigo or str(codigo).strip() == "":
                self.avisos.append(
                    f"Produto '{produto.get('denominacao', '?')}': sem 'codigo', "
                    f"não pode ser usado em PUT (atualização). Será gerado como POST."
                )
                # Gerar como POST
                item_post = self.gerar_json_post([produto])
                if item_post:
                    resultado.extend(item_post)
                continue

            item = {}
            item["seq"] = seq
            item["codigo"] = int(codigo) if isinstance(codigo, (int, float)) else codigo
            item["descricao"] = produto.get("descricao", "")
            item["denominacao"] = produto.get("denominacao", "")
            item["cpfCnpjRaiz"] = produto.get("cpfCnpjRaiz", "")
            situacao = produto.get("situacao", "ATIVADO")
            item["situacao"] = self.normalizar_situacao(situacao)
            item["modalidade"] = produto.get("modalidade", "")
            item["ncm"] = produto.get("ncm", "")

            item["atributos"] = produto.get("atributos", [])
            item["atributosMultivalorados"] = produto.get("atributosMultivalorados", [])
            item["atributosCompostos"] = produto.get("atributosCompostos", [])
            item["atributosCompostosMultivalorados"] = produto.get("atributosCompostosMultivalorados", [])
            item["codigosInterno"] = produto.get("codigosInterno", [])

            resultado.append(item)

        return resultado

    def gerar_json_api_post(self, produtos: list) -> list:
        """
        Gera JSON para POST via API nova: POST /ext/produto/{cpfCnpjRaiz}
        Usa o schema ProdutoIntegracaoRequestDTO.
        NÃO inclui: seq, cpfCnpjRaiz, situacao, codigo, versao (cpfCnpjRaiz vai na URL).
        """
        resultado = []
        for produto in produtos:
            item = {}
            item["descricao"] = produto.get("descricao", "")
            item["denominacao"] = produto.get("denominacao", "")
            item["modalidade"] = produto.get("modalidade", "")
            item["ncm"] = produto.get("ncm", "")

            # Atributos
            item["atributos"] = produto.get("atributos", [])
            item["atributosMultivalorados"] = produto.get("atributosMultivalorados", [])
            item["atributosCompostos"] = produto.get("atributosCompostos", [])
            item["atributosCompostosMultivalorados"] = produto.get("atributosCompostosMultivalorados", [])

            # Códigos internos
            item["codigosInterno"] = produto.get("codigosInterno", [])

            resultado.append(item)

        return resultado

    def gerar_json_api_put(self, produtos: list) -> list:
        """
        Gera JSON para PUT via API nova:
        - Nova versão: PUT /ext/produto/{cpfCnpjRaiz}/{codigo}
        - Retificação: PUT /ext/produto/{cpfCnpjRaiz}/{codigo}/{versao}
        Usa o schema ProdutoIntegracaoRequestDTO (mesmo do POST).
        NÃO inclui: seq, cpfCnpjRaiz, situacao, codigo, versao no body
        (codigo e versao vão na URL).
        """
        resultado = []
        for produto in produtos:
            codigo = produto.get("codigo")
            if not codigo or str(codigo).strip() == "":
                self.avisos.append(
                    f"Produto '{produto.get('denominacao', '?')}': sem 'codigo', "
                    f"será gerado como inclusão (POST). Para atualizar via API, "
                    f"use PUT /ext/produto/{{cpfCnpjRaiz}}/{{codigo}} com o código na URL."
                )

            item = {}
            item["descricao"] = produto.get("descricao", "")
            item["denominacao"] = produto.get("denominacao", "")
            item["modalidade"] = produto.get("modalidade", "")
            item["ncm"] = produto.get("ncm", "")

            item["atributos"] = produto.get("atributos", [])
            item["atributosMultivalorados"] = produto.get("atributosMultivalorados", [])
            item["atributosCompostos"] = produto.get("atributosCompostos", [])
            item["atributosCompostosMultivalorados"] = produto.get("atributosCompostosMultivalorados", [])
            item["codigosInterno"] = produto.get("codigosInterno", [])

            resultado.append(item)

        return resultado

    def gerar_json_completo(self, produtos: list) -> list:
        """
        Gera JSON no formato completo de exportação (como o portal exporta),
        incluindo seq, codigo, versao. Ordem dos campos idêntica ao portal.
        """
        resultado = []
        for seq, produto in enumerate(produtos, 1):
            item = {}
            # Ordem exata do portal: seq, codigo, descricao, denominacao, 
            # cpfCnpjRaiz, situacao, modalidade, ncm, versao, atributos,
            # atributosMultivalorados, atributosCompostos, 
            # atributosCompostosMultivalorados, codigosInterno
            item["seq"] = seq

            codigo = produto.get("codigo", seq)
            if isinstance(codigo, str) and codigo.strip() == "":
                codigo = seq
            try:
                codigo = int(codigo)
            except (ValueError, TypeError):
                codigo = seq
            item["codigo"] = codigo

            item["descricao"] = produto.get("descricao", "")
            item["denominacao"] = produto.get("denominacao", "")
            item["cpfCnpjRaiz"] = produto.get("cpfCnpjRaiz", "")
            situacao = produto.get("situacao", "ATIVADO")
            item["situacao"] = self.normalizar_situacao(situacao)
            item["modalidade"] = produto.get("modalidade", "")
            item["ncm"] = produto.get("ncm", "")
            # versao: só incluir se veio da planilha/JSON original (campo do servidor)
            versao = produto.get("versao")
            if versao and str(versao).strip():
                item["versao"] = str(versao).strip()
            item["atributos"] = produto.get("atributos", [])
            item["atributosMultivalorados"] = produto.get("atributosMultivalorados", [])
            item["atributosCompostos"] = produto.get("atributosCompostos", [])
            item["atributosCompostosMultivalorados"] = produto.get("atributosCompostosMultivalorados", [])
            item["codigosInterno"] = produto.get("codigosInterno", [])

            resultado.append(item)

        return resultado

    # ========================================================================
    # GERAÇÃO DE PLANILHA MODELO
    # ========================================================================

    def gerar_planilha_modelo(self, caminho_saida: str, atributos_extras: list = None):
        """
        Gera planilha modelo (.xlsx) com as colunas corretas e instruções.
        
        Args:
            caminho_saida: Caminho do arquivo .xlsx de saída
            atributos_extras: Lista de códigos ATT_ adicionais para incluir
        """
        wb = openpyxl.Workbook()

        # ---- Aba principal: PRODUTOS ----
        ws = wb.active
        ws.title = "PRODUTOS"

        # Estilos
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        att_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        att_multi_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        att_font = Font(color="FFFFFF", bold=True, size=10)
        obrig_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Cabeçalhos dos campos principais
        cabecalhos = list(COLUNAS_PRINCIPAIS)

        # Adicionar atributos padrão se não especificados
        if atributos_extras is None:
            # Atributos comuns para produtos de importação
            atributos_extras = [
                "ATT_14540",   # Condição do Produto
                "ATT_14545",   # País de Origem
                "ATT_14546",   # Validade do Produto
                "ATT_14547",   # Produto Controlado
                "ATT_14551",   # Registro ANVISA
                "ATT_14554",   # Produto Perigoso
                "ATT_14555",   # Fabricante/Exportador
                "ATT_14860",   # Nome Comercial
                "ATT_15120",   # Composição/Material
                "ATT_15121",   # Modelo/Referência
            ]

        # Adicionar colunas de atributos simples
        for att in atributos_extras:
            label = ATRIBUTOS_LABELS.get(att, "")
            cabecalhos.append(f"{att}" if not label else f"{att}")

        # Adicionar coluna de atributo multivalorado
        cabecalhos.append("ATT_14556_MULTI")

        # Escrever cabeçalhos
        for col_idx, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_idx, value=cab)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if cab.startswith("ATT_") and "_MULTI" in cab:
                cell.fill = att_multi_fill
                cell.font = att_font
            elif cab.startswith("ATT_"):
                cell.fill = att_fill
                cell.font = att_font
            else:
                cell.fill = header_fill
                cell.font = header_font

        # Linha 2: Descrições/labels dos cabeçalhos
        descricoes_row = 2
        descricoes = {
            "codigo": "Código do produto\n(vazio p/ novo,\npreencher p/ atualizar)",
            "denominacao": "Nome do produto\n(OBRIGATÓRIO)\nMáx 120 caracteres",
            "descricao": "Descrição detalhada\n(OBRIGATÓRIO)\nMáx 2000 caracteres",
            "cpfCnpjRaiz": "Seu CNPJ raiz 8 dígitos\n(OBRIGATÓRIO)\nSó números",
            "situacao": "ATIVADO ou DESATIVADO\n(padrão: ATIVADO)",
            "modalidade": "IMPORTACAO ou\nEXPORTACAO\n(OBRIGATÓRIO)",
            "ncm": "Código NCM 8 dígitos\n(OBRIGATÓRIO)\nSó números",
            "codigosInterno": "Códigos internos\nseparados por ;\n(opcional)",
        }
        for col_idx, cab in enumerate(cabecalhos, 1):
            if cab in descricoes:
                desc = descricoes[cab]
            elif cab.startswith("ATT_"):
                codigo_att = re.match(r"(ATT_\d+)", cab).group(1)
                desc = ATRIBUTOS_LABELS.get(codigo_att, "Atributo")
                if "_MULTI" in cab:
                    desc += "\n(valores separados por ;)"
            else:
                desc = ""

            cell = ws.cell(row=descricoes_row, column=col_idx, value=desc)
            cell.font = Font(italic=True, size=9, color="555555")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = border

        # Linha 3: Exemplo
        exemplo_row = 3
        exemplos = {
            "codigo": "",
            "denominacao": "ARCO NITI 12 (M) INF/SUP 10UN",
            "descricao": "Indicado Para Tratamentos Ortodônticos...",
            "cpfCnpjRaiz": "12345678",
            "situacao": "ATIVADO",
            "modalidade": "IMPORTACAO",
            "ncm": "90211010",
            "codigosInterno": "6128",
        }
        exemplos_att = {
            "ATT_14540": "01",
            "ATT_14545": "82",
            "ATT_14546": "INDETERMINADA",
            "ATT_14547": "false",
            "ATT_14551": "80853390002",
            "ATT_14554": "false",
            "ATT_14555": "HANGZHOU XINGCHEN 3B DENTAL INSTRUMENT E MATERIAL CO.,LTD",
            "ATT_14860": "ARCO NITI 12 (M) INF/SUP 10UN",
            "ATT_15120": "NIQUEL TITANIO",
            "ATT_15121": "AW002-12U",
        }
        for col_idx, cab in enumerate(cabecalhos, 1):
            if cab in exemplos:
                val = exemplos[cab]
            elif cab.startswith("ATT_"):
                codigo_att = re.match(r"(ATT_\d+)", cab).group(1)
                val = exemplos_att.get(codigo_att, "")
                if "_MULTI" in cab and codigo_att == "ATT_14556":
                    val = "11"
            else:
                val = ""

            cell = ws.cell(row=exemplo_row, column=col_idx, value=val)
            cell.font = Font(size=10, color="0070C0")
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border

        # Ajustar larguras
        larguras = {
            "codigo": 12,
            "denominacao": 45,
            "descricao": 60,
            "cpfCnpjRaiz": 15,
            "situacao": 14,
            "modalidade": 16,
            "ncm": 12,
            "codigosInterno": 22,
        }
        for col_idx, cab in enumerate(cabecalhos, 1):
            if cab in larguras:
                ws.column_dimensions[get_column_letter(col_idx)].width = larguras[cab]
            elif cab.startswith("ATT_"):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Altura das linhas
        ws.row_dimensions[1].height = 35
        ws.row_dimensions[2].height = 55
        ws.row_dimensions[3].height = 25

        # Congelar painel
        ws.freeze_panes = "A4"

        # ---- Aba de instruções ----
        ws_inst = wb.create_sheet("INSTRUÇÕES")
        instrucoes = [
            ["INSTRUÇÕES DE PREENCHIMENTO - CATÁLOGO DE PRODUTOS SISCOMEX"],
            [""],
            ["CAMPOS OBRIGATÓRIOS:"],
            ["  • denominacao: Nome do produto (máx 200 caracteres)"],
            ["  • descricao: Descrição detalhada do produto (máx 2000 caracteres)"],
            ["  • cpfCnpjRaiz: CNPJ raiz da empresa (8 dígitos, somente números)"],
            ["  • modalidade: IMPORTACAO ou EXPORTACAO (sem acento)"],
            ["  • ncm: Código NCM com 8 dígitos (somente números, sem pontos)"],
            [""],
            ["CAMPOS OPCIONAIS:"],
            ["  • codigo: Deixar VAZIO para novos produtos. Preencher para ATUALIZAR produto existente"],
            ["  • situacao: 'ATIVADO' (padrão) ou 'DESATIVADO'"],
            ["  • codigosInterno: Códigos internos da empresa, separados por ponto-e-vírgula (;)"],
            [""],
            ["ATRIBUTOS (colunas ATT_xxxxx):"],
            ["  • Os atributos disponíveis variam conforme o NCM do produto"],
            ["  • Atributos com sufixo _MULTI aceitam múltiplos valores separados por ;"],
            ["  • Consulte os atributos do NCM em: /cada/api/ext/atributo-ncm?ncm=XXXXXXXX&modalidade=IMPORTACAO"],
            [""],
            ["ATRIBUTOS COMUNS PARA PRODUTOS MÉDICOS/ODONTOLÓGICOS:"],
            [f"  • ATT_14540 - {ATRIBUTOS_LABELS.get('ATT_14540', '')} (01=Novo)"],
            [f"  • ATT_14545 - {ATRIBUTOS_LABELS.get('ATT_14545', '')} (82=China, 249=EUA, etc)"],
            [f"  • ATT_14546 - {ATRIBUTOS_LABELS.get('ATT_14546', '')} (INDETERMINADO, 2 ANOS, etc)"],
            [f"  • ATT_14547 - {ATRIBUTOS_LABELS.get('ATT_14547', '')} (true/false)"],
            [f"  • ATT_14551 - {ATRIBUTOS_LABELS.get('ATT_14551', '')} (número do registro)"],
            [f"  • ATT_14554 - {ATRIBUTOS_LABELS.get('ATT_14554', '')} (true/false)"],
            [f"  • ATT_14555 - {ATRIBUTOS_LABELS.get('ATT_14555', '')} (nome completo)"],
            [f"  • ATT_14556 - {ATRIBUTOS_LABELS.get('ATT_14556', '')} (MULTIVALORADO - códigos separados por ;)"],
            [f"  • ATT_14860 - {ATRIBUTOS_LABELS.get('ATT_14860', '')}"],
            [f"  • ATT_15120 - {ATRIBUTOS_LABELS.get('ATT_15120', '')}"],
            [f"  • ATT_15121 - {ATRIBUTOS_LABELS.get('ATT_15121', '')}"],
            [""],
            ["FORMATOS DE SAÍDA JSON:"],
            ["  1. API POST (incluir): Gera JSON ProdutoIntegracaoRequestDTO (sem seq/cpfCnpjRaiz/situacao)"],
            ["  2. API PUT (atualizar): Gera JSON ProdutoIntegracaoRequestDTO (para PUT /{cpfCnpjRaiz}/{codigo})"],
            ["  3. POST Lote (upload portal): Gera JSON ProdutoIntegracaoDTO com seq (endpoint depreciado)"],
            ["  4. PUT Lote (upload portal): Gera JSON ProdutoIntegracaoDTO com seq e codigo"],
            ["  5. Completo: Gera JSON igual ao exportado pelo portal (com seq/codigo/versao)"],
            [""],
            ["NOTAS:"],
            ["  • A linha 2 da aba PRODUTOS contém descrições dos campos (NÃO preencher)"],
            ["  • A linha 3 contém um exemplo (pode sobrescrever com seus dados)"],
            ["  • Dados começam na linha 4 em diante"],
            ["  • Linhas em branco serão ignoradas"],
        ]
        for row_idx, linha in enumerate(instrucoes, 1):
            cell = ws_inst.cell(row=row_idx, column=1, value=linha[0] if linha else "")
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="1F4E79")
            elif linha and linha[0].endswith(":"):
                cell.font = Font(bold=True, size=11)
            else:
                cell.font = Font(size=10)

        ws_inst.column_dimensions['A'].width = 100

        # ---- Aba de códigos de país ----
        ws_paises = wb.create_sheet("CÓDIGOS PAÍS")
        paises_comuns = [
            ["Código", "País"],
            ["82", "China"],
            ["249", "Estados Unidos"],
            ["105", "Alemanha"],
            ["399", "Japão"],
            ["Se precisar de mais códigos, consulte a tabela de países do Siscomex", ""],
        ]
        for row_idx, dados in enumerate(paises_comuns, 1):
            for col_idx, val in enumerate(dados, 1):
                cell = ws_paises.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

        ws_paises.column_dimensions['A'].width = 60
        ws_paises.column_dimensions['B'].width = 30

        wb.save(caminho_saida)
        print(f"\n📝 Planilha modelo salva em: {caminho_saida}")

    # ========================================================================
    # IMPORTAR JSON EXISTENTE PARA PLANILHA
    # ========================================================================

    def json_para_planilha(self, caminho_json: str, caminho_excel: str):
        """
        Converte um JSON exportado do portal para planilha Excel.
        Útil para editar produtos existentes e re-importar.
        """
        print(f"\n📂 Lendo JSON: {caminho_json}")

        with open(caminho_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

        if not isinstance(dados, list):
            dados = [dados]

        print(f"   {len(dados)} produtos encontrados no JSON")

        # Coletar todos os códigos de atributos usados
        todos_att_simples = set()
        todos_att_multi = set()

        for produto in dados:
            for att in produto.get("atributos", []):
                todos_att_simples.add(att["atributo"])
            for att in produto.get("atributosMultivalorados", []):
                todos_att_multi.add(att["atributo"])

        todos_att_simples = sorted(todos_att_simples)
        todos_att_multi = sorted(todos_att_multi)

        # Criar planilha
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PRODUTOS"

        # Estilos
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        att_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        att_multi_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        att_font = Font(color="FFFFFF", bold=True, size=10)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Montar cabeçalhos
        cabecalhos = list(COLUNAS_PRINCIPAIS)
        cabecalhos.extend(todos_att_simples)
        cabecalhos.extend([f"{att}_MULTI" for att in todos_att_multi])

        # Escrever cabeçalhos
        for col_idx, cab in enumerate(cabecalhos, 1):
            cell = ws.cell(row=1, column=col_idx, value=cab)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if "_MULTI" in cab:
                cell.fill = att_multi_fill
                cell.font = att_font
            elif cab.startswith("ATT_"):
                cell.fill = att_fill
                cell.font = att_font
            else:
                cell.fill = header_fill
                cell.font = header_font

        # Escrever dados
        for row_idx, produto in enumerate(dados, 2):
            for col_idx, cab in enumerate(cabecalhos, 1):
                valor = ""
                if cab in COLUNAS_PRINCIPAIS:
                    valor_raw = produto.get(cab, "")
                    if cab == "codigosInterno":
                        codigos = produto.get("codigosInterno", [])
                        valor = ";".join(codigos) if codigos else ""
                    else:
                        valor = str(valor_raw) if valor_raw is not None else ""
                elif "_MULTI" in cab:
                    codigo_att = re.match(r"(ATT_\d+)", cab).group(1)
                    for att in produto.get("atributosMultivalorados", []):
                        if att["atributo"] == codigo_att:
                            valor = ";".join(att.get("valores", []))
                            break
                elif cab.startswith("ATT_"):
                    for att in produto.get("atributos", []):
                        if att["atributo"] == cab:
                            valor = att.get("valor", "")
                            break

                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # Ajustar larguras
        for col_idx, cab in enumerate(cabecalhos, 1):
            if cab == "descricao":
                ws.column_dimensions[get_column_letter(col_idx)].width = 60
            elif cab == "denominacao":
                ws.column_dimensions[get_column_letter(col_idx)].width = 45
            elif cab.startswith("ATT_"):
                ws.column_dimensions[get_column_letter(col_idx)].width = 25
            else:
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        wb.save(caminho_excel)
        print(f"\n✅ Planilha salva em: {caminho_excel}")
        print(f"   {len(dados)} produtos exportados para a planilha.")

    # ========================================================================
    # MÉTODO PRINCIPAL DE CONVERSÃO
    # ========================================================================

    def converter(self, caminho_excel: str, caminho_json_saida: str = None,
                  modo: str = "post", indent: int = 2) -> str:
        """
        Método principal: converte planilha Excel em JSON.
        
        Args:
            caminho_excel: Caminho do arquivo .xlsx de entrada
            caminho_json_saida: Caminho do arquivo .json de saída (auto-gerado se None)
            modo: 'post' (criar), 'put' (atualizar) ou 'completo' (formato exportação)
            indent: Indentação do JSON (2 para legível, None para compacto)
        
        Returns:
            Caminho do arquivo JSON gerado
        """
        self.erros = []
        self.avisos = []

        # Ler planilha
        produtos = self.ler_planilha(caminho_excel)

        # Verificar erros
        if self.erros:
            print(f"\n❌ {len(self.erros)} ERRO(S) ENCONTRADO(S):")
            for erro in self.erros:
                print(f"   ⛔ {erro}")
            print("\n⚠️  Corrija os erros acima e tente novamente.")
            return None

        if not produtos:
            print("\n⚠️  Nenhum produto encontrado na planilha.")
            return None

        # Mostrar avisos
        if self.avisos:
            print(f"\n⚠️  {len(self.avisos)} AVISO(S):")
            for aviso in self.avisos:
                print(f"   ⚡ {aviso}")

        # Gerar JSON conforme modo
        modo = modo.lower()
        if modo == "post":
            json_data = self.gerar_json_post(produtos)
            sufixo = "_POST"
        elif modo == "put":
            json_data = self.gerar_json_put(produtos)
            sufixo = "_PUT"
        elif modo == "api_post":
            json_data = self.gerar_json_api_post(produtos)
            sufixo = "_API_POST"
        elif modo == "api_put":
            json_data = self.gerar_json_api_put(produtos)
            sufixo = "_API_PUT"
        elif modo == "completo":
            json_data = self.gerar_json_completo(produtos)
            sufixo = "_COMPLETO"
        else:
            print(f"\n❌ Modo '{modo}' inválido. Use: post, put, api_post, api_put ou completo")
            return None

        # Verificar avisos pós-geração
        if self.avisos:
            for aviso in self.avisos:
                if aviso not in [a for a in self.avisos[:len(self.avisos)//2]]:
                    print(f"   ⚡ {aviso}")

        # Determinar caminho de saída
        if caminho_json_saida is None:
            base = os.path.splitext(caminho_excel)[0]
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            caminho_json_saida = f"{base}{sufixo}_{timestamp}.json"

        # Salvar JSON
        with open(caminho_json_saida, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=indent)

        tamanho_kb = os.path.getsize(caminho_json_saida) / 1024
        print(f"\n{'='*70}")
        print(f"✅ JSON GERADO COM SUCESSO!")
        print(f"   📄 Arquivo: {caminho_json_saida}")
        print(f"   📦 Produtos: {len(json_data)}")
        print(f"   📏 Tamanho: {tamanho_kb:.1f} KB")
        print(f"   🔧 Modo: {modo.upper()}")
        print(f"{'='*70}")

        return caminho_json_saida


# ============================================================================
# INTERFACE DE LINHA DE COMANDO (CLI)
# ============================================================================

def exibir_menu():
    """Exibe menu principal interativo."""
    print("\n" + "=" * 70)
    print("  CONVERSOR CATÁLOGO DE PRODUTOS - PORTAL ÚNICO SISCOMEX (CATP)")
    print("=" * 70)
    print()
    print("  Escolha uma opção:")
    print()
    print("  === API NOVA (ProdutoIntegracaoRequestDTO) ===")
    print("  [1] 📡 Excel → JSON (API POST - incluir produto)")
    print("  [2] 📡 Excel → JSON (API PUT - nova versão/retificação)")
    print()
    print("  === UPLOAD EM LOTE NO PORTAL (ProdutoIntegracaoDTO) ===")
    print("  [3] 📊 Excel → JSON (POST lote - criar novos produtos)")
    print("  [4] 📊 Excel → JSON (PUT lote - atualizar produtos)")
    print()
    print("  === UTILITÁRIOS ===")
    print("  [5] 📊 Excel → JSON (Completo - formato exportação do portal)")
    print("  [6] 📝 Gerar planilha modelo (.xlsx)")
    print("  [7] 📥 JSON → Excel (importar JSON do portal para planilha)")
    print("  [8] ✅ Validar planilha (sem gerar JSON)")
    print("  [0] ❌ Sair")
    print()
    return input("  Opção: ").strip()


def solicitar_caminho(mensagem: str, extensao: str = None, deve_existir: bool = True) -> str:
    """Solicita um caminho de arquivo ao usuário."""
    while True:
        caminho = input(f"\n  {mensagem}: ").strip()
        if not caminho:
            print("  ⚠️  Caminho não pode ser vazio.")
            continue

        # Remover aspas se colocadas
        caminho = caminho.strip('"').strip("'")

        if deve_existir and not os.path.exists(caminho):
            print(f"  ⚠️  Arquivo não encontrado: {caminho}")
            continue

        if extensao:
            if not caminho.lower().endswith(extensao.lower()):
                caminho += extensao

        return caminho


def main():
    """Função principal - modo interativo."""
    conversor = ConversorCatalogoSiscomex()

    # Se argumentos de linha de comando foram passados
    if len(sys.argv) > 1:
        # Modo CLI direto
        import argparse
        parser = argparse.ArgumentParser(
            description="Conversor Excel → JSON para API CATP Siscomex"
        )
        parser.add_argument("arquivo", help="Caminho do arquivo Excel (.xlsx)")
        parser.add_argument(
            "-m", "--modo",
            choices=["post", "put", "api_post", "api_put", "completo"],
            default="api_post",
            help="Modo de geração: api_post (padrão), api_put, post (lote), put (lote) ou completo"
        )
        parser.add_argument(
            "-o", "--output",
            help="Caminho do arquivo JSON de saída"
        )
        parser.add_argument(
            "--modelo",
            action="store_true",
            help="Gerar planilha modelo"
        )
        parser.add_argument(
            "--json-para-excel",
            help="Converter JSON do portal para Excel"
        )
        parser.add_argument(
            "--compacto",
            action="store_true",
            help="Gerar JSON compacto (sem indentação)"
        )

        args = parser.parse_args()

        if args.modelo:
            conversor.gerar_planilha_modelo(args.arquivo)
        elif args.json_para_excel:
            conversor.json_para_planilha(args.json_para_excel, args.arquivo)
        else:
            indent = None if args.compacto else 2
            conversor.converter(args.arquivo, args.output, args.modo, indent)
        return

    # Modo interativo
    while True:
        opcao = exibir_menu()

        if opcao == "0":
            print("\n  👋 Até logo!")
            break

        elif opcao in ["1", "2", "3", "4", "5"]:
            modos = {
                "1": "api_post",
                "2": "api_put",
                "3": "post",
                "4": "put",
                "5": "completo"
            }
            modo = modos[opcao]

            caminho_excel = solicitar_caminho(
                "Caminho da planilha Excel (.xlsx)",
                extensao=".xlsx",
                deve_existir=True
            )

            resultado = conversor.converter(caminho_excel, modo=modo)

            if resultado:
                print(f"\n  ✅ Arquivo JSON pronto para uso na API!")
                if modo == "api_post":
                    print(f"  📡 Endpoint: POST /catp/api/ext/produto/{{cpfCnpjRaiz}}")
                    print(f"  ℹ️  cpfCnpjRaiz vai na URL, NÃO no body do JSON")
                elif modo == "api_put":
                    print(f"  📡 Endpoint: PUT /catp/api/ext/produto/{{cpfCnpjRaiz}}/{{codigo}}")
                    print(f"  ℹ️  cpfCnpjRaiz e codigo vão na URL, NÃO no body do JSON")
                elif modo == "post":
                    print(f"  📡 Endpoint: POST /catp/api/ext/produto (lote/upload portal)")
                elif modo == "put":
                    print(f"  📡 Endpoint: POST /catp/api/ext/produto (lote/upload portal)")

        elif opcao == "6":
            caminho_saida = solicitar_caminho(
                "Caminho para salvar a planilha modelo (.xlsx)",
                extensao=".xlsx",
                deve_existir=False
            )
            conversor.gerar_planilha_modelo(caminho_saida)
            print(f"\n  ✅ Planilha modelo criada! Preencha e use opção 1-5.")

        elif opcao == "7":
            caminho_json = solicitar_caminho(
                "Caminho do arquivo JSON exportado do portal",
                extensao=".json",
                deve_existir=True
            )
            caminho_excel = solicitar_caminho(
                "Caminho para salvar a planilha Excel (.xlsx)",
                extensao=".xlsx",
                deve_existir=False
            )
            conversor.json_para_planilha(caminho_json, caminho_excel)

        elif opcao == "8":
            caminho_excel = solicitar_caminho(
                "Caminho da planilha Excel (.xlsx)",
                extensao=".xlsx",
                deve_existir=True
            )
            conversor.erros = []
            conversor.avisos = []
            produtos = conversor.ler_planilha(caminho_excel)

            if conversor.erros:
                print(f"\n  ❌ {len(conversor.erros)} ERRO(S):")
                for erro in conversor.erros:
                    print(f"     ⛔ {erro}")
            else:
                print(f"\n  ✅ Planilha válida! {len(produtos)} produtos prontos.")

            if conversor.avisos:
                print(f"\n  ⚠️  {len(conversor.avisos)} AVISO(S):")
                for aviso in conversor.avisos:
                    print(f"     ⚡ {aviso}")

        else:
            print("\n  ⚠️  Opção inválida. Tente novamente.")

        input("\n  Pressione ENTER para continuar...")


if __name__ == "__main__":
    main()
