# -*- coding: utf-8 -*-
"""
Aplicação Web - Conversor Catálogo de Produtos Siscomex (CATP API)
Flask + Interface moderna
"""

import json
import os
import sys
import io
import uuid
import tempfile
import shutil
import zipfile
from datetime import datetime

from flask import (
    Flask, render_template, request, send_file,
    jsonify, redirect, url_for, flash, session
)
from werkzeug.utils import secure_filename

# Importar conversor: tenta local primeiro (deploy), depois diretório pai (dev)
try:
    from conversor_catalogo_siscomex import ConversorCatalogoSiscomex, ATRIBUTOS_LABELS
except ImportError:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from conversor_catalogo_siscomex import ConversorCatalogoSiscomex, ATRIBUTOS_LABELS

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'conversor-siscomex-catp-2026-secret')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Diretório temporário para uploads/downloads
UPLOAD_FOLDER = os.path.join(tempfile.gettempdir(), 'siscomex_catp_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EXTENSOES_PERMITIDAS_EXCEL = {'.xlsx', '.xls'}
EXTENSOES_PERMITIDAS_JSON = {'.json'}

# ============================================================================
# CARREGAR ATRIBUTOS VÁLIDOS POR NCM (arquivo oficial do Siscomex)
# ============================================================================
ATRIBUTOS_POR_NCM = {}  # { "90211010": { "ATT_14545": {...}, ... } }

def carregar_atributos_ncm():
    """Carrega o JSON oficial de atributos por NCM do Siscomex."""
    global ATRIBUTOS_POR_NCM
    # Tenta encontrar o arquivo em vários caminhos possíveis
    caminhos = [
        os.path.join(os.path.dirname(__file__), 'ATRIBUTOS_POR_NCM.json'),
        os.path.join(os.path.dirname(os.path.dirname(__file__)), 'ATRIBUTOS_POR_NCM_2026_02_22.json'),
    ]
    for caminho in caminhos:
        if os.path.exists(caminho):
            try:
                with open(caminho, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                for ncm_entry in data.get('listaNcm', []):
                    ncm_code = ncm_entry['codigoNcm'].replace('.', '')
                    attrs = {}
                    for att in ncm_entry.get('listaAtributos', []):
                        attrs[att['codigo']] = {
                            'obrigatorio': att.get('obrigatorio', False),
                            'multivalorado': att.get('multivalorado', False),
                            'modalidade': att.get('modalidade', ''),
                        }
                    ATRIBUTOS_POR_NCM[ncm_code] = attrs
                print(f"[CATP] Carregados atributos para {len(ATRIBUTOS_POR_NCM)} NCMs de {caminho}")
                return
            except Exception as e:
                print(f"[CATP] Erro ao carregar atributos: {e}")
    print("[CATP] AVISO: Arquivo de atributos por NCM não encontrado. Validação de atributos desabilitada.")

# Carregar ao iniciar
carregar_atributos_ncm()


def filtrar_atributos_por_ncm(produtos, avisos_extra=None):
    """
    Filtra atributos de cada produto para manter APENAS os válidos para o NCM.
    Remove atributos que não existem na lista oficial e avisa.
    Também injeta obrigatórios faltantes quando possível.
    """
    if not ATRIBUTOS_POR_NCM:
        return  # Sem dados, não filtra
    
    if avisos_extra is None:
        avisos_extra = []
    
    for i, produto in enumerate(produtos):
        ncm = produto.get('ncm', '').strip()
        if not ncm or ncm not in ATRIBUTOS_POR_NCM:
            continue
        
        validos = ATRIBUTOS_POR_NCM[ncm]
        nome = produto.get('denominacao', f'Produto {i+1}')[:50]
        
        # Filtrar atributos simples
        atributos_orig = produto.get('atributos', [])
        atributos_filtrados = []
        removidos = []
        for att in atributos_orig:
            cod = att.get('atributo', '')
            if cod in validos:
                atributos_filtrados.append(att)
            else:
                removidos.append(cod)
        
        if removidos:
            avisos_extra.append(
                f"Produto '{nome}': Removidos atributos não válidos para NCM {ncm}: {', '.join(removidos)}"
            )
        produto['atributos'] = atributos_filtrados
        
        # Filtrar atributos multivalorados
        multi_orig = produto.get('atributosMultivalorados', [])
        multi_filtrados = []
        removidos_multi = []
        for att in multi_orig:
            cod = att.get('atributo', '')
            if cod in validos and validos[cod].get('multivalorado'):
                multi_filtrados.append(att)
            elif cod in validos and not validos[cod].get('multivalorado'):
                # Atributo existe mas não é multivalorado - converter para simples
                valores = att.get('valores', [])
                if valores:
                    atributos_filtrados.append({
                        'atributo': cod,
                        'valor': valores[0]
                    })
                    avisos_extra.append(
                        f"Produto '{nome}': {cod} convertido de multivalorado para simples"
                    )
            else:
                removidos_multi.append(cod)
        
        if removidos_multi:
            avisos_extra.append(
                f"Produto '{nome}': Removidos atributos multi não válidos para NCM {ncm}: {', '.join(removidos_multi)}"
            )
        produto['atributosMultivalorados'] = multi_filtrados
        produto['atributos'] = atributos_filtrados
        
        # Verificar obrigatórios faltantes
        existentes = {a.get('atributo') for a in produto['atributos']}
        existentes_multi = {a.get('atributo') for a in produto['atributosMultivalorados']}
        
        for cod, info in validos.items():
            if info['obrigatorio'] and cod not in existentes and cod not in existentes_multi:
                avisos_extra.append(
                    f"Produto '{nome}': FALTA atributo obrigatório {cod} para NCM {ncm}!"
                )


def extensao_permitida(filename, permitidas):
    return os.path.splitext(filename)[1].lower() in permitidas


def converter_xls_para_xlsx(caminho_xls: str) -> str:
    """Converte arquivo .xls (formato antigo) para .xlsx usando xlrd + openpyxl."""
    import xlrd
    from openpyxl import Workbook

    # Ler o .xls
    wb_xls = xlrd.open_workbook(caminho_xls)
    ws_xls = wb_xls.sheet_by_index(0)

    # Criar .xlsx
    wb_xlsx = Workbook()
    ws_xlsx = wb_xlsx.active

    for row in range(ws_xls.nrows):
        for col in range(ws_xls.ncols):
            cell = ws_xls.cell(row, col)
            valor = cell.value
            # xlrd retorna floats para números inteiros — corrigir
            if cell.ctype == xlrd.XL_CELL_NUMBER and valor == int(valor):
                valor = int(valor)
            ws_xlsx.cell(row=row + 1, column=col + 1, value=valor)

    caminho_xlsx = caminho_xls.rsplit('.', 1)[0] + '.xlsx'
    wb_xlsx.save(caminho_xlsx)
    wb_xlsx.close()
    wb_xls.release_resources()

    return caminho_xlsx


def limpar_arquivos_antigos():
    """Remove arquivos temporários com mais de 1 hora."""
    agora = datetime.now().timestamp()
    try:
        for f in os.listdir(UPLOAD_FOLDER):
            caminho = os.path.join(UPLOAD_FOLDER, f)
            if os.path.isfile(caminho):
                idade = agora - os.path.getmtime(caminho)
                if idade > 3600:  # 1 hora
                    os.remove(caminho)
    except Exception:
        pass


# ============================================================================
# ROTAS PRINCIPAIS
# ============================================================================

@app.route('/')
def index():
    """Página principal."""
    limpar_arquivos_antigos()
    return render_template('index.html')


@app.route('/converter', methods=['POST'])
def converter():
    """Converte Excel para JSON."""
    if 'arquivo' not in request.files:
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}), 400

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo selecionado.'}), 400

    if not extensao_permitida(arquivo.filename, EXTENSOES_PERMITIDAS_EXCEL):
        ext = os.path.splitext(arquivo.filename)[1].lower()
        if ext == '.xls':
            return jsonify({'sucesso': False, 'erro': 'O formato antigo .xls não é suportado. Abra o arquivo no Excel e salve como .xlsx (Pasta de Trabalho do Excel).'}), 400
        return jsonify({'sucesso': False, 'erro': 'Formato inválido. Envie um arquivo .xlsx'}), 400

    modo = request.form.get('modo', 'post')
    if modo not in ['post', 'put', 'api_post', 'api_put', 'completo']:
        return jsonify({'sucesso': False, 'erro': 'Modo inválido.'}), 400

    # Valores padrão para colunas que podem não existir na planilha
    cnpj_padrao = request.form.get('cnpj_padrao', '').strip()
    modalidade_padrao = request.form.get('modalidade_padrao', '').strip()
    pais_origem_padrao = request.form.get('pais_origem_padrao', '').strip()
    validade_padrao = request.form.get('validade_padrao', '').strip()
    controlado_padrao = request.form.get('controlado_padrao', '').strip()
    perigoso_padrao = request.form.get('perigoso_padrao', '').strip()
    fabricante_padrao = request.form.get('fabricante_padrao', '').strip()
    embalagem_padrao = request.form.get('embalagem_padrao', '').strip()
    operador_estrangeiro = request.form.get('operador_estrangeiro', '').strip()

    try:
        # Salvar arquivo temporário
        uid = str(uuid.uuid4())[:8]
        nome_seguro = secure_filename(arquivo.filename)
        caminho_excel = os.path.join(UPLOAD_FOLDER, f"{uid}_{nome_seguro}")
        arquivo.save(caminho_excel)

        # Auto-converter .xls → .xlsx
        ext = os.path.splitext(caminho_excel)[1].lower()
        if ext == '.xls':
            try:
                caminho_xlsx = converter_xls_para_xlsx(caminho_excel)
                os.remove(caminho_excel)
                caminho_excel = caminho_xlsx
            except Exception as e:
                os.remove(caminho_excel)
                return jsonify({
                    'sucesso': False,
                    'erro': f'Erro ao converter .xls para .xlsx: {str(e)}. Tente abrir no Excel e salvar como .xlsx manualmente.'
                }), 400

        # Montar defaults
        defaults = {}
        if cnpj_padrao:
            defaults['cpfCnpjRaiz'] = cnpj_padrao
        if modalidade_padrao:
            defaults['modalidade'] = modalidade_padrao

        # Auto-truncar campos longos?
        auto_truncar = request.form.get('auto_truncar', 'false').lower() == 'true'

        # Converter
        conversor = ConversorCatalogoSiscomex(auto_truncar=auto_truncar)
        produtos = conversor.ler_planilha(caminho_excel, defaults=defaults)

        # Injetar atributos obrigatórios nos produtos que não os possuem
        for produto in produtos:
            atributos = produto.get('atributos', [])
            multi = produto.get('atributosMultivalorados', [])
            codigos_simples = {a.get('atributo') for a in atributos}
            codigos_multi = {a.get('atributo') for a in multi}

            # ATT_14545 — País de Origem (simples)
            if pais_origem_padrao and 'ATT_14545' not in codigos_simples:
                atributos.insert(0, {'atributo': 'ATT_14545', 'valor': pais_origem_padrao})

            # ATT_14546 — Validade (simples)
            if validade_padrao and 'ATT_14546' not in codigos_simples:
                atributos.append({'atributo': 'ATT_14546', 'valor': validade_padrao})

            # ATT_14547 — Controlado (simples)
            if controlado_padrao and 'ATT_14547' not in codigos_simples:
                atributos.append({'atributo': 'ATT_14547', 'valor': controlado_padrao})

            # ATT_14554 — Perigoso (simples)
            if perigoso_padrao and 'ATT_14554' not in codigos_simples:
                atributos.append({'atributo': 'ATT_14554', 'valor': perigoso_padrao})

            # ATT_14555 — Fabricante/Exportador (simples)
            if fabricante_padrao and 'ATT_14555' not in codigos_simples:
                atributos.append({'atributo': 'ATT_14555', 'valor': fabricante_padrao})

            # ATT_14556 — Embalagem (MULTIVALORADO)
            if embalagem_padrao and 'ATT_14556' not in codigos_multi:
                multi.append({'atributo': 'ATT_14556', 'valores': [embalagem_padrao]})

            produto['atributos'] = atributos
            produto['atributosMultivalorados'] = multi

            # codigoOperadorEstrangeiro
            if operador_estrangeiro and not produto.get('codigoOperadorEstrangeiro'):
                produto['codigoOperadorEstrangeiro'] = operador_estrangeiro

        if conversor.erros:
            # Limpar
            os.remove(caminho_excel)
            return jsonify({
                'sucesso': False,
                'erro': 'Erros encontrados na planilha.',
                'erros': conversor.erros,
                'avisos': conversor.avisos
            }), 400

        if not produtos:
            os.remove(caminho_excel)
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum produto encontrado na planilha. Verifique se os dados começam na linha correta.'
            }), 400

        # Gerar JSON
        if modo == 'post':
            json_data = conversor.gerar_json_post(produtos)
        elif modo == 'put':
            json_data = conversor.gerar_json_put(produtos)
        elif modo == 'api_post':
            json_data = conversor.gerar_json_api_post(produtos)
        elif modo == 'api_put':
            json_data = conversor.gerar_json_api_put(produtos)
        else:
            json_data = conversor.gerar_json_completo(produtos)

        # Salvar JSON temporário para download
        nome_json = f"{uid}_CATALOGO_{modo.upper()}.json"
        caminho_json = os.path.join(UPLOAD_FOLDER, nome_json)
        with open(caminho_json, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

        # Limpar Excel
        os.remove(caminho_excel)

        # Resposta
        json_preview = json.dumps(json_data[:3], ensure_ascii=False, indent=2)
        if len(json_data) > 3:
            json_preview += f"\n\n... e mais {len(json_data) - 3} produto(s)"

        return jsonify({
            'sucesso': True,
            'mensagem': f'{len(json_data)} produto(s) convertido(s) com sucesso!',
            'total_produtos': len(json_data),
            'modo': modo.upper(),
            'arquivo_download': nome_json,
            'preview': json_preview,
            'json_completo': json.dumps(json_data, ensure_ascii=False, indent=2),
            'avisos': conversor.avisos
        })

    except zipfile.BadZipFile:
        return jsonify({
            'sucesso': False,
            'erro': 'O arquivo não é um .xlsx válido. Provavelmente está no formato antigo .xls renomeado para .xlsx. Abra o arquivo no Excel, clique em Salvar Como e escolha "Pasta de Trabalho do Excel (.xlsx)".'
        }), 400
    except Exception as e:
        return jsonify({
            'sucesso': False,
            'erro': f'Erro inesperado: {str(e)}'
        }), 500


@app.route('/json-para-excel', methods=['POST'])
def json_para_excel():
    """Converte JSON do portal para Excel."""
    if 'arquivo' not in request.files:
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}), 400

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo selecionado.'}), 400

    if not extensao_permitida(arquivo.filename, EXTENSOES_PERMITIDAS_JSON):
        return jsonify({'sucesso': False, 'erro': 'Formato inválido. Envie um arquivo .json'}), 400

    try:
        uid = str(uuid.uuid4())[:8]
        nome_seguro = secure_filename(arquivo.filename)
        caminho_json = os.path.join(UPLOAD_FOLDER, f"{uid}_{nome_seguro}")
        arquivo.save(caminho_json)

        nome_excel = f"{uid}_CATALOGO_EDITAVEL.xlsx"
        caminho_excel = os.path.join(UPLOAD_FOLDER, nome_excel)

        conversor = ConversorCatalogoSiscomex()
        conversor.json_para_planilha(caminho_json, caminho_excel)

        # Contar produtos
        with open(caminho_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)
            total = len(dados) if isinstance(dados, list) else 1

        os.remove(caminho_json)

        return jsonify({
            'sucesso': True,
            'mensagem': f'{total} produto(s) convertido(s) para Excel!',
            'total_produtos': total,
            'arquivo_download': nome_excel
        })

    except json.JSONDecodeError:
        return jsonify({'sucesso': False, 'erro': 'JSON inválido. Verifique o formato do arquivo.'}), 400
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': f'Erro: {str(e)}'}), 500


@app.route('/modelo')
def baixar_modelo():
    """Gera e baixa a planilha modelo."""
    try:
        uid = str(uuid.uuid4())[:8]
        nome = f"MODELO_catalogo_siscomex.xlsx"
        caminho = os.path.join(UPLOAD_FOLDER, f"{uid}_{nome}")

        conversor = ConversorCatalogoSiscomex()
        conversor.gerar_planilha_modelo(caminho)

        return send_file(
            caminho,
            as_attachment=True,
            download_name=nome,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@app.route('/download/<nome_arquivo>')
def download(nome_arquivo):
    """Download de arquivos gerados."""
    nome_seguro = secure_filename(nome_arquivo)
    caminho = os.path.join(UPLOAD_FOLDER, nome_seguro)

    if not os.path.exists(caminho):
        return jsonify({'sucesso': False, 'erro': 'Arquivo não encontrado ou expirado.'}), 404

    # Determinar nome de download amigável
    partes = nome_seguro.split('_', 1)
    nome_download = partes[1] if len(partes) > 1 else nome_seguro

    mimetype = 'application/json'
    if nome_seguro.endswith('.xlsx'):
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return send_file(
        caminho,
        as_attachment=True,
        download_name=nome_download,
        mimetype=mimetype
    )


@app.route('/validar', methods=['POST'])
def validar():
    """Valida planilha sem gerar JSON."""
    if 'arquivo' not in request.files:
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}), 400

    arquivo = request.files['arquivo']
    if not extensao_permitida(arquivo.filename, EXTENSOES_PERMITIDAS_EXCEL):
        ext = os.path.splitext(arquivo.filename)[1].lower()
        if ext == '.xls':
            return jsonify({'sucesso': False, 'erro': 'O formato antigo .xls não é suportado. Abra o arquivo no Excel e salve como .xlsx.'}), 400
        return jsonify({'sucesso': False, 'erro': 'Formato inválido. Envie .xlsx'}), 400

    try:
        uid = str(uuid.uuid4())[:8]
        caminho = os.path.join(UPLOAD_FOLDER, f"{uid}_{secure_filename(arquivo.filename)}")
        arquivo.save(caminho)

        conversor = ConversorCatalogoSiscomex()
        produtos = conversor.ler_planilha(caminho)

        os.remove(caminho)

        if conversor.erros:
            return jsonify({
                'sucesso': False,
                'valido': False,
                'total_produtos': 0,
                'erros': conversor.erros,
                'avisos': conversor.avisos
            })

        return jsonify({
            'sucesso': True,
            'valido': True,
            'total_produtos': len(produtos),
            'erros': [],
            'avisos': conversor.avisos
        })

    except Exception as e:
        return jsonify({'sucesso': False, 'erro': str(e)}), 500


@app.route('/atributos')
def listar_atributos():
    """Retorna lista de atributos conhecidos."""
    return jsonify(ATRIBUTOS_LABELS)


@app.route('/converter-operador', methods=['POST'])
def converter_operador():
    """Converte Excel de operadores estrangeiros para JSON."""
    if 'arquivo' not in request.files:
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}), 400

    arquivo = request.files['arquivo']
    if arquivo.filename == '':
        return jsonify({'sucesso': False, 'erro': 'Nenhum arquivo selecionado.'}), 400

    ext = os.path.splitext(arquivo.filename)[1].lower()
    if ext not in {'.xlsx', '.xls'}:
        return jsonify({'sucesso': False, 'erro': 'Formato inválido. Envie um arquivo .xlsx ou .xls'}), 400

    cnpj_raiz = request.form.get('cnpj_raiz_operador', '').strip()

    try:
        uid = str(uuid.uuid4())[:8]
        nome_seguro = secure_filename(arquivo.filename)
        caminho_excel = os.path.join(UPLOAD_FOLDER, f"{uid}_{nome_seguro}")
        arquivo.save(caminho_excel)

        # Auto-converter .xls → .xlsx
        if ext == '.xls':
            try:
                caminho_xlsx = converter_xls_para_xlsx(caminho_excel)
                os.remove(caminho_excel)
                caminho_excel = caminho_xlsx
            except Exception as e:
                os.remove(caminho_excel)
                return jsonify({
                    'sucesso': False,
                    'erro': f'Erro ao converter .xls: {str(e)}'
                }), 400

        import openpyxl
        wb = openpyxl.load_workbook(caminho_excel, read_only=True, data_only=True)
        ws = wb.active

        # Ler cabeçalhos (primeira linha)
        cabecalhos = []
        for cell in ws[1]:
            val = str(cell.value).strip() if cell.value else ''
            cabecalhos.append(val.lower())

        # Mapear colunas conhecidas
        MAPA_COLUNAS = {
            'nome': ['nome', 'name', 'razao social', 'razão social', 'empresa', 'company'],
            'logradouro': ['logradouro', 'endereco', 'endereço', 'address', 'rua', 'street'],
            'numero': ['numero', 'número', 'nro', 'num', 'number', 'no'],
            'complemento': ['complemento', 'complement', 'comp', 'apto', 'sala'],
            'codigoPais': ['codigopais', 'codigo pais', 'código país', 'pais', 'país', 'country', 'country code'],
            'nomeCidade': ['cidade', 'city', 'nomecidade', 'nome cidade', 'municipio', 'município'],
            'estado': ['estado', 'state', 'uf', 'provincia', 'província', 'province'],
            'codigoPostal': ['cep', 'codigopostal', 'codigo postal', 'código postal', 'zip', 'zipcode', 'zip code', 'postal code', 'postal'],
            'telefone': ['telefone', 'phone', 'tel', 'fone', 'telephone'],
            'email': ['email', 'e-mail', 'mail'],
            'cpfCnpjRaiz': ['cnpj', 'cpfcnpjraiz', 'cpf cnpj raiz', 'cnpj raiz', 'cpf/cnpj'],
        }

        mapa = {}  # campo_json -> indice_coluna
        for campo, aliases in MAPA_COLUNAS.items():
            for idx, cab in enumerate(cabecalhos):
                if cab in aliases or cab.replace('_', ' ') in aliases:
                    mapa[campo] = idx
                    break

        # Ler linhas de dados
        operadores = []
        avisos = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            def get_val(campo):
                idx = mapa.get(campo)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    val = str(row[idx]).strip()
                    if val.endswith('.0'):
                        try:
                            float(val)
                            val = val[:-2]
                        except ValueError:
                            pass
                    return val
                return ''

            nome = get_val('nome')
            if not nome:
                avisos.append(f'Linha {row_num}: sem nome, ignorada.')
                continue

            operador = {}

            # cpfCnpjRaiz: da planilha ou do formulário
            cpf_cnpj = get_val('cpfCnpjRaiz')
            if not cpf_cnpj and cnpj_raiz:
                cpf_cnpj = cnpj_raiz
            if cpf_cnpj:
                cpf_cnpj = cpf_cnpj.replace('.', '').replace('-', '').replace('/', '').replace(' ', '')
                operador['cpfCnpjRaiz'] = cpf_cnpj

            operador['nome'] = nome

            # Endereço
            endereco = {}
            logradouro = get_val('logradouro')
            if logradouro:
                endereco['logradouro'] = logradouro
            numero = get_val('numero')
            if numero:
                endereco['numero'] = numero
            complemento = get_val('complemento')
            if complemento:
                endereco['complemento'] = complemento
            if endereco:
                operador['endereco'] = endereco

            # País
            cod_pais = get_val('codigoPais')
            if cod_pais:
                operador['codigoPais'] = cod_pais

            # Cidade
            cidade = get_val('nomeCidade')
            if cidade:
                operador['nomeCidade'] = cidade

            # Estado
            estado = get_val('estado')
            if estado:
                operador['estado'] = estado

            # CEP
            cep = get_val('codigoPostal')
            if cep:
                operador['codigoPostal'] = cep

            # Telefone
            telefone = get_val('telefone')
            if telefone:
                operador['telefone'] = telefone

            # Email
            email = get_val('email')
            if email:
                operador['email'] = email

            operadores.append(operador)

        wb.close()
        os.remove(caminho_excel)

        if not operadores:
            return jsonify({
                'sucesso': False,
                'erro': 'Nenhum operador encontrado na planilha. Verifique se há uma coluna "nome" com dados.',
                'avisos': avisos
            }), 400

        # Salvar JSON para download
        nome_json = f"{uid}_OPERADORES_ESTRANGEIROS.json"
        caminho_json = os.path.join(UPLOAD_FOLDER, nome_json)
        with open(caminho_json, 'w', encoding='utf-8') as f:
            json.dump(operadores, f, ensure_ascii=False, indent=2)

        # Preview
        json_preview = json.dumps(operadores[:3], ensure_ascii=False, indent=2)
        if len(operadores) > 3:
            json_preview += f"\n\n... e mais {len(operadores) - 3} operador(es)"

        return jsonify({
            'sucesso': True,
            'mensagem': f'{len(operadores)} operador(es) convertido(s) com sucesso!',
            'total': len(operadores),
            'arquivo_download': nome_json,
            'preview': json_preview,
            'json_completo': json.dumps(operadores, ensure_ascii=False, indent=2),
            'avisos': avisos
        })

    except Exception as e:
        return jsonify({'sucesso': False, 'erro': f'Erro inesperado: {str(e)}'}), 500


# ============================================================================
# INICIALIZAÇÃO
# ============================================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'true').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug)
